import logging, traceback
import base64
import csv
import os
from nfty.barcodes import BarcodePDF417
from io import StringIO, BytesIO
from PIL import Image
from openpyxl import load_workbook
from nfty.docs import Document

logger = logging.getLogger("AppLogger")

def parse_csv(file, delim=",", quotechar='"', binary=False):
    if binary:
        with StringIO(file.read().decode("utf-8")) as x:
            csvfile = csv.reader(x, delimiter=delim, quotechar=quotechar)
            for row in csvfile:
                yield row
    else:
        with open(file, "r") as f:
            csvfile = csv.reader(f, delimiter=delim, quotechar=quotechar)
            for row in csvfile:
                yield row


def write_csv(filename, data, header=None, delim=",", quotechar='"', binary=True):
    writemode = "w"
    if binary:
        writemode = "wb"
    with open(filename, writemode, encoding="utf-8") as f:
        csvfile = csv.writer(
            f,
            delimiter=delim,
            quotechar=quotechar,
            quoting=csv.QUOTE_MINIMAL,
            lineterminator="\n",
        )
        if header:
            csvfile.writerow(header)
        csvfile.writerows(data)
    return True


class Excel:
    def __init__(self, file, binary=True):
        self.file = file
        self.header = []
        if binary:
            self._binary()
        else:
            self._notbinary()

    def _notbinary(self):
        self.wb = load_workbook(filename=self.file, read_only=True)

    def _binary(self):
        self.wb = load_workbook(filename=BytesIO(self.file.read()))

    def ws(self, sheetname=False):
        if not sheetname:
            ws = self.wb[self.wb.sheetnames[0]]
        else:
            ws = self.wb[sheetname]
        xlist = []
        xa = xlist.append
        for row in ws.rows:
            xa([cell.value for cell in row])
        return xlist

    def wsgen(self, sheetname=False):
        if not sheetname:
            ws = self.wb[self.wb.sheetnames[0]]
        else:
            ws = self.wb[sheetname]
        for row in ws.rows:
            yield [cell.value for cell in row]


class FileUpload:
    def __init__(self, request, client):
        self.c = client
        self.request = request
        self.files = request.files
        self.header = []
        self.columns = {}
        self.filename = None
        self.file = None
        self.count = 0
        self.response = {}

    def license(self):
        for self.filename, self.file in list(self.files.items()):
            if x := BarcodePDF417(self.file, save=True).detect():
                return x
        return {}

    def process(self):
        for self.filename, self.file in list(self.files.items()):
            self.detect()
        return self.response

    # def onboard(self, ws):
    #     """
    #     This takes in an Excel file and with the parent id values, loads users into the system.
    #     """
    #     parentids = ("id", "clientid", "parentid", "parent")
    #     parent = constants.CLIENT_ID_DEFAULT_API
    #     for line in ws:
    #         meta = {self.columns[i]: v for i, v in enumerate(line) if v}
    #         for p in parentids:
    #             if meta.get(p):
    #                 parent = meta.pop(p)
    #                 break
    #         if not parent:
    #             parent = self.c.id
    #         if meta:
    #             self.count += 1
    #             e = Enroll()
    #             e.post(meta, parent=parent, skip=True)
    #             e.set_status(status=constants.ONBOARDING_SEND_LINK)
    #     self.response = f"Successfully imported {self.count} clients"
    #     return self

    def detect(self):
        if self.filename == "file":
            self.response = Document(self.request, self.c).new()
            return self
        name, ext = self.filename.rsplit(".", 1)
        # if ext in ("xls", "xlsx"):
        #     ws = Excel(self.file.file).wsgen()
        #     header = ws.__next__()
        #     if len(fileTypes["enroll"].intersection(set(header))) == 6:
        #         self.columns = {i: v.lower() for i, v in enumerate(header)}
        #         self.onboard(ws)
        #     elif len(fileTypes["enroll_2"].intersection(set(header))) == 6:
        #         rename = fileTypes["enroll_map"]
        #         header = [rename.get(h, h) for h in header]
        #         self.columns = {i: v.lower() for i, v in enumerate(header)}
        #         self.onboard(ws)
        # else:
        self.response = Document(self.request, self.c).upload(self.file)
        return self

    def logo(self):
        def resize_image(image, dimensions):
            i = Image.open(image.file)
            i.thumbnail(dimensions)
            return i

        imagedict = {}
        name, ext = os.path.splitext(self.filename)
        image = resize_image(self.file, (400, 400))
        buffer = BytesIO()
        image.save(buffer, optimize=True, format="PNG")
        imagedict[name + ".PNG"] = base64.b64encode(buffer.getvalue()).decode()
        self.response.update(imagedict)
        return self



def delete_file(file_path):
    try:
        os.remove(file_path)
    except:
        pass

